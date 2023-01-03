# ------------------------------------------------------------------------------------------------
# -- coding                                   | utf-8
# -- Author                                   | Sergei Chernyahovsky
# -- Site                                     | http://sergeicher.pro/
# -- Favorite Quote                           | “Always code as if the guy who ends up
#                                                   maintaining your code will be a violent
#                                                       psychopath who knows where you live”
# -- Language                                 | Python
# -- Version                                  | 3.11
# -- WebDriver                                | Selenium
# -- Version                                  | 4.6.0
# -- Description                              | Excel External files
# Usage Example
# Import this module and use these functions
# ------------------------------------------------------------------------------------------------


# import openpyxl module
import openpyxl
import allure
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


class Excel:

    # Create workbook object
    # Get workbook active sheet object from the active attribute Cell objects also
    # have a row, column, and coordinate attributes that provide
    # location information for the cell.
    @staticmethod
    @allure.step("Load workbook object")
    def LoadWorkbookObject(filePath: str):
        try:
            excelObj = openpyxl.load_workbook(filePath)
            sheetObj = excelObj.active
            return sheetObj
        except FileExistsError:
            print(f"File {filePath} doesn't exists")

    '''
    Reading from Spreadsheets, Multiple Cells
    Give the location of the file to open the workbook, workbook object is created
    Get workbook active sheet object from the active attribute Cell objects
    We can get the count of the total rows and columns using the
    max_row and max_column respectively. We can use these values inside the
    for loop to get the value of the desired row or column or any cell depending upon
    the situation. Let’s see how to get the value of the first column and first row.
    '''

    @staticmethod
    @allure.step("Read column from excel file")
    def ReadColumn(filePath: str, col: int) -> list:
        row = Excel.LoadWorkbookObject(filePath=filePath).max_row
        result = []
        for i in range(1, row + 1):
            cellObj = Excel.LoadWorkbookObject(filePath=filePath).cell(row=i, column=col)
            result.append(cellObj.value)
        return result[1:]

    @staticmethod
    @allure.step("Read row from excel file")
    def ReadRow(filePath: str, row: int) -> list:
        column = Excel.LoadWorkbookObject(filePath=filePath).max_column
        result = []
        for i in range(1, column + 1):
            cellObj = Excel.LoadWorkbookObject(filePath=filePath).cell(row=row, column=i)
            result.append(cellObj.value)
        return result

    @staticmethod
    @allure.step("Read active sheet from excel file")
    def ReadSpreadsheets(filePath: str):
        cellObj = Excel.LoadWorkbookObject(filePath=filePath)
        result = []
        for i in cellObj:
            for j in i:
                result.append(j.value)
        return result

    @staticmethod
    @allure.step("Read value from specific cell in excel file")
    def ReadCellsValue(filePath: str, rowInt: int, colInt: int) -> str:
        cellObj = Excel.LoadWorkbookObject(filePath=filePath).cell(row=rowInt, column=colInt)
        return cellObj.value

    '''
    Writing to Spreadsheets
    '''

    # REWRITE !!! or Creates a new file
    @staticmethod
    @allure.step("Saves value to excel file, if file exists -> rewrites")
    def SaveValueNewXL(filePath: str, val: str):
        excelObj = openpyxl.Workbook()
        sheetObj = excelObj.active
        val1 = sheetObj.cell(row=1, column=1)
        val1.value = val
        excelObj.save(filePath)

    # File must be created first
    # CELL Example : path, "A1" or "B6", "ABCD"
    @staticmethod
    @allure.step("Update value in specific cell in excel file")
    def UpdateValueInCell(filePath: str, cell: str, value: str):
        try:
            excelObj = load_workbook(filename=filePath)
            sheetObj = excelObj.active
            sheetObj[cell] = value
            excelObj.save(filename=filePath)
        except FileExistsError:
            print(f"File {filePath} doesn't exists")

        # finally:
        #     print(print(f"File {filePath} doesn't exist, \n Message: "))
        #         # print(f"File {filePath} doesn't exist, \n Message: ", m)

    '''Adding Images
    For the purpose of importing images inside our worksheet,
     we would be using openpyxl.drawing.image.Image. The method is a wrapper over
     PIL.Image method found in PIL (pillow) library. Due to which it is necessary
     for the PIL (pillow) library to be installed in order to use this method.
    Image Used:'''

    @staticmethod
    @allure.step("Saves image in new excel file")
    def SaveImageInNewXL(filePath: str, imagePath: str, cell: str):
        wb = openpyxl.Workbook()
        sheet = wb.active
        img = Image(imagePath)
        sheet.add_image(img, cell)
        wb.save(filePath)

    @staticmethod
    @allure.step("Saves image in existing excel file")
    def SaveImage(filePath: str, imagePath: str, cell: str):
        try:
            wb = openpyxl.load_workbook(filePath)
            sheet = wb.active
            img = Image(imagePath)
            sheet.add_image(img, cell)
            wb.save(filePath)
        except FileExistsError:
            print(f"File {filePath} doesn't exists")

############################################################################

# continue from here implementing as needed


############################################################################
